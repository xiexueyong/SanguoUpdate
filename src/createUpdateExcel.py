#! /usr/bin/env python
# -*- coding: utf-8 -*-

import md5
import os
import time

from openpyxl import Workbook
from openpyxl import load_workbook
from GlobalConfig import * 
from Md5Util import * 

#添加一条数据，为一个channel，添加一个oldtag到targetTag，
global updateitemStartHead
updateitemStartHead = 2
def processOneTagOnOneChannelId(fileNameSuffix,channelId,oldTag,targetTag,ws):
    fn = 'update-%s-%s-%s.zip'%(oldTag,targetTag,fileNameSuffix)
    fileurl = preFtpURL+fn
    ws.cell('A%d'%updateitemStartHead).value = idHead
    ws.cell('B%d'%updateitemStartHead).value = channelId
    ws.cell('C%d'%updateitemStartHead).value = oldTag
    ws.cell('D%d'%updateitemStartHead).value = targetTag
    ws.cell('E%d'%updateitemStartHead).value = fileurl
    ws.cell('F%d'%updateitemStartHead).value = getMD5(win_updateFilePath+'/'+fn)
    ws.cell('G%d'%updateitemStartHead).value = time.strftime("%Y-%m-%d %H:%M:%S")
    print fileurl
    print updateitemStartHead
    
    global updateitemStartHead
    updateitemStartHead = updateitemStartHead + 1
    global idHead
    idHead = idHead+1
    
#添加一个channel的 所有oldtag到targetTag的数据  
def processOneChannelId(fileNameSuffix,channelId,oldTags,targetTag,ws):
    for tag in oldTags:
        processOneTagOnOneChannelId(fileNameSuffix,channelId,tag,targetTag,ws)
#针对一个更新文件，为所有共同使用这个更新文件的渠道 设置数据       
def processOneUpdatefile(file_channelids,fileNameSuffix,oldTags,targetTag,ws):
    print fileNameSuffix
    print file_channelids
    for id in file_channelids:
        processOneChannelId(fileNameSuffix,id,oldTags,targetTag,ws)
#出事换update.xlsx文件头
def initialiseUpdateExcel(ws):
    ws.cell('A1').value = 'id'
    ws.cell('B1').value = "channel_id"
    ws.cell('C1').value = 'curversion'
    ws.cell('D1').value = 'targetTag'
    ws.cell('E1').value = 'url'
    ws.cell('F1').value = 'md5'
    ws.cell('G1').value = 'created_at'
#生成excel，针对所有的更新文件，为所有渠道设置数据
def createUpdateExcel():
    updateExcel = win_updateFilePath+'/core_update.xlsx'
    if os.path.exists(updateExcel):
        wb = load_workbook(updateExcel)
        ws = wb.get_sheet_by_name('update') 
        global updateitemStartHead
        updateitemStartHead = len(ws.rows)+1
        global idHead
        idHead = ws.cell(row = len(ws.rows), column = 1).value+1
        print len(ws.rows)
    else:
        wb = Workbook(encoding='utf-8')
        ws = wb.active
        ws.title= 'update'  
    
    initialiseUpdateExcel(ws)
    for relation in channels:
        processOneUpdatefile(useRelation[relation],relation,oldTags,targetTag,ws)
    wb.save(updateExcel)
if __name__ == '__main__':
    createUpdateExcel()


    
        
    
    
