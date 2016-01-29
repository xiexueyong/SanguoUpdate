#! /usr/bin/env python
# -*- coding: utf-8 -*-

import md5
import os
import time

from openpyxl import Workbook
from openpyxl import load_workbook
from GlobalConfig import * 
from Md5Util import * 

upgradeitemCount = 2
def processOneChannelAPK(channelId,version,apkurl,md5,ws):
    ws.cell('A%d'%upgradeitemCount).value = upgradeitemCount+idHead
    ws.cell('B%d'%upgradeitemCount).value = channelId
    ws.cell('C%d'%upgradeitemCount).value = version
    ws.cell('D%d'%upgradeitemCount).value = apkurl
    ws.cell('E%d'%upgradeitemCount).value = md5
    ws.cell('F%d'%upgradeitemCount).value = time.strftime("%Y-%m-%d %H:%M:%S")
    #print apkurl
    
    global upgradeitemCount
    upgradeitemCount = upgradeitemCount + 1
    
     
def initialiseUpgradeExcel(ws):
    ws.cell('A1').value = 'id'
    ws.cell('B1').value = "channel_id"
    ws.cell('C1').value = 'version'
    ws.cell('D1').value = 'url'
    ws.cell('E1').value = 'md5'
    ws.cell('F1').value = 'created_at'
    
def createUpgradeExcel():
    channelwb = load_workbook(win_updateFilePath+'/core_channel.xlsx')
    channelws = channelwb.get_sheet_by_name("channel")
    
    wb = Workbook(encoding='utf-8')
    ws = wb.active
    ws.title= 'upgrade'
    initialiseUpgradeExcel(ws)
    wb.save(win_updateFilePath+'/core_upgrade.xlsx') 
    
    for i in range(2,channelws.max_row+1):
        channelName = channelws.cell(row = i, column = 3).value
        channelId = channelws.cell(row = i, column = 1).value
        fn = 'Binglintianxia_Release_151217_v2.0.0b_%s.apk'%(channelName)
        localurl = ur'E:/Binglintianxia2.0.0/'+fn
        ftpurl = apkPreFtpURL+fn  
        #print channelName
        #print channelId
        #print ftpurl
        if not os.path.exists(localurl):
            print fn
            continue
        md5 = getMD5(localurl) 
       
        processOneChannelAPK(channelId,'v2.0.0b',ftpurl,md5,ws)
        
    wb.save(win_updateFilePath+'/core_upgrade.xlsx')
    print win_updateFilePath         
createUpgradeExcel()