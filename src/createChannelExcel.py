#! /usr/bin/env python
# -*- coding: utf-8 -*-

import md5
import os
import time

from openpyxl import Workbook
from openpyxl import load_workbook
from GlobalConfig import *
from Md5Util import * 


    
def initialiseChannelVersionExcel(ws):
    ws.cell('A1').value = 'id'
    ws.cell('B1').value = "title"
    ws.cell('C1').value = 'slug'
    ws.cell('D1').value = 'version'
def modifyChannelVersion(channelID,version,ws):
    for i in range(2,ws.max_row+1):
        version_e = ws.cell(row = i, column = 4)
        id_e = ws.cell(row = i, column = 1)
        #print id_e.value,version_e.value,"    ",channelID,version
        if id_e.value == channelID or str(id_e.value) == channelID:
            print channelID
            version_e.value = version
    
def createChannelVersionExcel():
    wb = load_workbook(win_updateFilePath+'/core_channel.xlsx')
    ws = wb.active
    ws.title= 'channel'
    initialiseChannelVersionExcel(ws)
    print ws.max_row
    print ws.max_column
    #for i range(0,ws.max_row):
    for channelsFileName in channels:
        for channel in useRelation[channelsFileName]:
            print 'channel in excel : '+str(channel)
            modifyChannelVersion(str(channel),targetTag,ws)
    wb.save(win_updateFilePath+'/core_channel.xlsx')
if __name__ == '__main__':        
    createChannelVersionExcel()  

    
    
