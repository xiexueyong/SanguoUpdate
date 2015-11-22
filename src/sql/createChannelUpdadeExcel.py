# -*- coding:utf-8 -*-
#!/bin/usr/python

#120.132.91.101
import openpyxl
from  openpyxl import load_workbook
from  openpyxl import worksheet
from  openpyxl import workbook
from ConvertExcelSqlTool import *
import cgt

def createSQLFile(excelurl,excel):    
    readexcel = load_workbook(excelurl+'.xlsx')
    sheetname =  readexcel.get_sheet_names()[0] #readexcel.get_sheet_by_name(cgt.excelname)
    re = readexcel.get_sheet_by_name(sheetname)
    heightrow = re.max_row
    lengthcolumn = re.max_column
    
    for i in range(cgt.startrow, heightrow+1):
        DSHcontent = []
        for j in range(cgt.startcolumn,lengthcolumn+1):
            DSHcontent.append(re.cell(row=i,column=j).value)
        #print 2222,DSHcontent #type(DSHcontent[2])==type('')
        if excel == 'core_update':
            sqlWriteHandleUpDate(DSHcontent,excelurl+'.sql',excel)
        elif excel == 'core_channel':
            sqlWriteHandleChannel(DSHcontent,excelurl+'.sql',excel)
        elif excel == 'core_upgrade':
            sqlWriteHandleUpGrade(DSHcontent,excelurl+'.sql',excel)

for excel in cgt.excelnames:
    createSQLFile(cgt.excelpath+excel,excel)
